"""Import Western's registrar course export (Excel) and upsert to Supabase.

The registrar (Andrew Pocock, Registrar Information Systems) sends a full course
listing as an .xlsx each term. This is the same public catalog data the calendar
scraper collects, but complete (all subjects, all affiliate campuses) and with a
few extra fields: antirequisites, "extra information" (lecture hours), and campus
flags.

It upserts on `code`, the same key scrape_courses.py uses, so re-running either
tool merges cleanly rather than duplicating.

Usage:
    python import_courses_xlsx.py --file "course export.xlsx" --dry-run
    python import_courses_xlsx.py --file "course export.xlsx" --dry-run --limit 20
    python import_courses_xlsx.py --file "course export.xlsx"          # write to DB

Reads SUPABASE_URL / SUPABASE_SECRET_KEY from ../.env.local (or env vars), the
same as scrape_courses.py.
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from dataclasses import asdict, dataclass
from pathlib import Path

import openpyxl
import requests
from dotenv import load_dotenv

# Column order in the "courses(2)" sheet of the registrar export.
COL_SUBJECT = 2       # COURSE CODE SUBJECT     e.g. "COMPSCI"
COL_CATALOG = 3       # COURSE CODE CATALOG NUMBER  e.g. "1027A"
COL_TITLE = 4         # COURSE TITLE
COL_DESC = 5          # COURSE DESCRIPTION
COL_PREREQ = 6        # PREREQUISITES
COL_ANTIREQ = 7       # ANTIREQUISITES
COL_CREDIT = 8        # CREDIT WEIGHT
COL_EXTRA = 9         # EXTRA INFORMATION
COL_MAIN = 10         # MAIN     (0/1)
COL_HURON = 11        # HURON    (0/1)
COL_BRESCIA = 12      # BRESCIA  (0/1)
COL_KINGS = 13        # KINGS    (0/1)

# (column index, display name) for the four campus flags.
CAMPUS_COLS = [
    (COL_MAIN, "Main"),
    (COL_HURON, "Huron"),
    (COL_BRESCIA, "Brescia"),
    (COL_KINGS, "King's"),
]

# Leading labels the registrar bakes into the free-text fields. We strip them so
# the stored value matches what scrape_courses.py produces (bare text, no label).
LABEL_RE = re.compile(
    r"^\s*(Prerequisite\(s\)|Antirequisite\(s\)|Extra Information)\s*:\s*",
    re.IGNORECASE,
)


@dataclass
class Course:
    code: str                    # "COMPSCI 1027A"
    title: str
    description: str | None
    prereqs: str | None
    antireqs: str | None
    extra_info: str | None
    campuses: list[str]
    credit_weight: float | None


def clean(value) -> str | None:
    """Trim, strip a leading label, and collapse empties to None."""
    if value is None:
        return None
    text = str(value).strip()
    text = LABEL_RE.sub("", text).strip()
    return text or None


def campuses_for(row) -> list[str]:
    return [name for col, name in CAMPUS_COLS if row[col] == 1]


def parse_workbook(path: Path) -> list[Course]:
    """Read every data row, then merge duplicate codes into one course each.

    ~700 codes appear on more than one row — they're campus/prereq variants of
    the same course. We keep one row per code (preferring the Main-campus row,
    which carries the general prerequisites) and union the campus flags across
    all of that code's rows.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["courses(2)"] if "courses(2)" in wb.sheetnames else wb.active

    # code -> {"primary": row_tuple, "campuses": set[str], "is_main": bool}
    merged: dict[str, dict] = {}
    order: list[str] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        subject = clean(row[COL_SUBJECT])
        catalog = clean(row[COL_CATALOG])
        if not subject or not catalog:
            continue
        code = f"{subject} {catalog}"
        is_main = row[COL_MAIN] == 1

        if code not in merged:
            merged[code] = {"primary": row, "is_main": is_main, "campuses": set()}
            order.append(code)
        else:
            # Prefer a Main-campus row as the primary record; the affiliate
            # rows often add campus-registration caveats to the prereqs.
            if is_main and not merged[code]["is_main"]:
                merged[code]["primary"] = row
                merged[code]["is_main"] = True
        merged[code]["campuses"].update(campuses_for(row))

    courses: list[Course] = []
    for code in order:
        row = merged[code]["primary"]
        credit = row[COL_CREDIT]
        campuses = [name for _, name in CAMPUS_COLS if name in merged[code]["campuses"]]
        courses.append(Course(
            code=code,
            title=clean(row[COL_TITLE]) or code,
            description=clean(row[COL_DESC]),
            prereqs=clean(row[COL_PREREQ]),
            antireqs=clean(row[COL_ANTIREQ]),
            extra_info=clean(row[COL_EXTRA]),
            campuses=campuses,
            credit_weight=float(credit) if isinstance(credit, (int, float)) else None,
        ))
    return courses


def upsert_to_supabase(courses: list[Course]) -> None:
    """Batch-upsert by `code` via PostgREST (mirrors scrape_courses.py)."""
    url = os.environ["SUPABASE_URL"].rstrip("/")
    key = os.environ["SUPABASE_SECRET_KEY"]
    endpoint = f"{url}/rest/v1/courses?on_conflict=code"
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates,return=minimal",
    }

    rows = [asdict(c) for c in courses]
    BATCH = 500
    for start in range(0, len(rows), BATCH):
        chunk = rows[start:start + BATCH]
        resp = requests.post(endpoint, headers=headers, json=chunk, timeout=60)
        if not resp.ok:
            raise RuntimeError(f"Upsert failed ({resp.status_code}): {resp.text}")
        print(f"  upserted {start + len(chunk)}/{len(rows)}", file=sys.stderr)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--file", required=True, help="Path to the registrar .xlsx export")
    parser.add_argument("--dry-run", action="store_true", help="Parse and print but don't write")
    parser.add_argument("--limit", type=int, help="In dry-run, print only the first N courses")
    args = parser.parse_args()

    env_path = Path(__file__).resolve().parent.parent / ".env.local"
    load_dotenv(env_path)
    if "SUPABASE_URL" not in os.environ and (v := os.environ.get("NEXT_PUBLIC_SUPABASE_URL")):
        os.environ["SUPABASE_URL"] = v

    courses = parse_workbook(Path(args.file))
    print(f"Parsed {len(courses)} unique courses from {args.file}.", file=sys.stderr)

    if args.dry_run:
        for c in (courses[:args.limit] if args.limit else courses):
            campuses = ",".join(c.campuses) or "-"
            print(f"  {c.code} | {c.title} | wt={c.credit_weight} | [{campuses}] | "
                  f"prereq={(c.prereqs or '')[:50]!r}")
        return

    if not courses:
        print("No courses to upsert.", file=sys.stderr)
        return

    if "SUPABASE_URL" not in os.environ or "SUPABASE_SECRET_KEY" not in os.environ:
        print("ERROR: SUPABASE_URL and SUPABASE_SECRET_KEY must be set "
              "(via .env.local or environment).", file=sys.stderr)
        sys.exit(1)

    upsert_to_supabase(courses)
    print("Done.", file=sys.stderr)


if __name__ == "__main__":
    main()
